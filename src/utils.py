import textwrap
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

from src.agent import Human_Agent
from src.patient import Patient
from src.constants import TASK_CONTENT, PROJECT_DESCRIPTION

def generate_combined_score_table(all_agents: list[Human_Agent]):
    """
    Generate a combined score table from multiple agents' evaluation data.
    This function processes a list of Human_Agent objects, each containing a DataFrame
    with evaluation scores for various models. It extracts the top 1 to top 4 models
    for each evaluation criterion (excluding the first column and overall score) and
    combines the results into a single DataFrame.
    Args:
        all_agents (list[Human_Agent]): A list of Human_Agent objects, each containing
                                        a DataFrame with evaluation scores.
    Returns:
        pd.DataFrame: A DataFrame containing the combined top models for each criterion
                      from all agents.
    """

    # Initialize a dictionary to store combined data
    combined_data = {}

    # Data processing for each agent
    for agent in all_agents:
        
        df = agent.table_df

        # Initialize a dictionary to store top models for each criterion
        top_models = {}

        # Iterate through each evaluation criterion (excluding the first column and overall score)
        for column in df.columns[1:-1]:
            # Sort models by the criterion scores in descending order
            sorted_models = df.sort_values(by=column, ascending=False).iloc[:, 0]
            
            if 'Model_' in sorted_models[0]:
                sorted_models = sorted_models.str.replace('Model_', '')
            elif 'Model ' in sorted_models[0]:
                sorted_models = sorted_models.str.replace('Model ', '')

            # Get the top 1 to top 4 models
            top_models[column] = list(sorted_models[:4])

        # Store the results in combined_data
        combined_data[agent.name] = top_models

    # Create a DataFrame for the final large table
    large_table = []

    # Prepare the final structure
    for agent, criteria_data in combined_data.items():
        row = {'Agent': agent}
        for criterion, models in criteria_data.items():
            for i, model in enumerate(models):
                row[f'{criterion}_Top{i+1}'] = model
        large_table.append(row)

    final_df = pd.DataFrame(large_table)

    return final_df

def save_with_merged_cells(final_df, output_path):
    """
    Save a DataFrame to an Excel file with merged cells for headers.

    Parameters:
    final_df (pd.DataFrame): The DataFrame to be saved to the Excel file.
    output_path (str): The file path where the Excel file will be saved.

    Returns:
    None
    """
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Prepare merged headers
    headers = ["Agent"]
    sub_headers = []
    for column in final_df.columns[1:]:
        if "_Top" in column:
            criterion = column.split("_Top")[0]
            if criterion not in headers:
                headers.append(criterion)
                sub_headers.append("Top1")
            else:
                if sub_headers[-1] == "Top1":
                    sub_headers.append("Top2")
                elif sub_headers[-1] == "Top2":
                    sub_headers.append("Top3")
                elif sub_headers[-1] == "Top3":
                    sub_headers.append("Top4")
                elif sub_headers[-1] == "Top4":
                    sub_headers.append("Top1")

    # Write headers and sub-headers with merged cells
    data_start_col = 2
    col_index = data_start_col
    for i, header in enumerate(headers):
        if i == 0:
            # Write the first header 'Agent' and merge the cells
            ws.cell(1, 1).value = header
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[ws.cell(1, 1).column_letter].width = 20
        else:
            # Write the rest of the headers and merge the cells
            span = len([h for h in sub_headers if h.startswith("Top")]) // len(headers[1:])
            ws.cell(1, col_index).value = header
            ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + span - 1)
            ws.cell(1, col_index).alignment = Alignment(horizontal="center", vertical="center")
            col_index += span

    # Write data in each single cell (Top1, Top2, etc.)
    for col_num, sub_header in enumerate(sub_headers, data_start_col):
        ws.cell(2, col_num).value = sub_header
        ws.cell(2, col_num).alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row in dataframe_to_rows(final_df, index=False, header=False):
        ws.append(row)
        for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in cell:
                if c.column != 1:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

    # Save the workbook
    wb.save(output_path)
    

def define_input_message(patient: Patient):
    """
    Constructs an input message string for a given patient by combining project description,
    patient clinical records, diagnostic reports from various models, and the known target illness.
    Args:
        patient (Patient): An instance of the Patient class containing clinical records,
                           diagnostic reports, and the target illness.
    Returns:
        str: A formatted string containing the project description, patient clinical records,
             diagnostic reports from various models, and the known target illness.
    """
    proj_content = textwrap.dedent(
        f"""
        {PROJECT_DESCRIPTION}
        患者信息是：{patient.clinical_records}
        """
        + "\n".join(
            f"{model}模型的诊断报告是: {report}"
            for model, report in patient.diagnostic_reports_.items()
        )
    )

    input_message = textwrap.dedent(
        f"""
        {proj_content},
        已知正确的诊断疾病结果是{patient.target_illness},
        {TASK_CONTENT},
        """
    )
    
    return input_message